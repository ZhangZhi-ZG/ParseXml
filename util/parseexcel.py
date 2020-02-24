import openpyxl
from openpyxl.styles import Alignment,PatternFill,Side,Border,Font
from openpyxl.utils import get_column_letter
def write_excel(results,statistics,filename):
    #创建excel表格对象
    mywb = openpyxl.Workbook()
    #创建指定sheet页
    mywb.create_sheet('Test Statistics', 0)
    mywb.create_sheet('Passed Test Cases Details',1)
    mywb.create_sheet('Failed Test Cases Details', 2)
    #使用指定sheet页
    mysheet_statistics = mywb.get_sheet_by_name("Test Statistics")
    mysheet_pass = mywb.get_sheet_by_name("Passed Test Cases Details")
    mysheet_fail = mywb.get_sheet_by_name("Failed Test Cases Details")
    #表头
    headers = ["Name","Tags","Status","Message","Elapsed"]
    #行
    row_statistics,row_fail,row_pass = 1,1,1
    # 列-65为大写字母“A”的ASCII码
    col = 65
    #添加第一行表头数据
    for header in headers:
        mysheet_pass[chr(col) + str(row_pass)] = header
        mysheet_fail[chr(col) + str(row_fail)] = header
        col += 1
    #对一个sheet页写入case统计信息数据
    for statistic in statistics:
        mysheet_statistics[chr(65) + str(row_statistics)] = statistic[0]
        mysheet_statistics[chr(66)+str(row_statistics)]= statistic[1]
        row_statistics += 1
    #添加case测试结果数据
    for result in results:
        if result[2] == "FAIL":
            row_fail +=1
            for i in range(len(result)):
                #列-65为大写字母“A”的ASCII码
                col = 65
                for data in result:
                    mysheet_fail[chr(col)+str(row_fail)]=data
                    col += 1
        else:
            row_pass +=1
            for j in range(len(result)):
                #列-65为大写字母“A”的ASCII码
                col = 65
                for data in result:
                    mysheet_pass[chr(col)+str(row_pass)]=data
                    col += 1
    #调用格式化函数对每个sheet页单元格进行格式化
    for sheet in mywb.worksheets:
        if sheet.max_row > 1:
            formatsetting(sheet)
        else:
            #删除空sheet页
            mywb.remove(sheet)
        #调整列宽，实现自适应
        col_auto_scal(sheet)
    mywb.save(filename)

#设置单元格对齐方式
def formatsetting(sheet):
    #设置单元格框线
    thin = Side(border_style="thin",color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    #获取sheet页的最大行数及列数
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_col+1):
            cell = sheet.cell(row=i,column=j)
            cell.border = border
            if max_col > 2:
                if i == 1:
                    #行首单元格居中处理
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    #行首单元格填充蓝色，其中"0000FF"为蓝色对应的十六进制颜色码
                    cell.fill = PatternFill(start_color="4169E1", end_color="4169E1",fill_type="solid")
                else:
                    if j in [1,2,4]:
                        #单元格左对齐
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif j == 3:
                        # 单元格居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.value == "PASS":
                            #case通过，单元格填充绿色
                            cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
                        else:
                            # case失败，单元格填充红色
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    else:
                        # 单元格居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                #所有单元格都居中处理
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if j == 1:
                    # 第一列单元格填充蓝色，其中"4169E1"为蓝色对应的十六进制颜色码
                    cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
                elif j == 2 and i == 1:
                    if cell.value == "FAIL":
                        # 字体格式为红色、宋体、不加粗
                        cell.font = Font(name='宋体', size=11, bold=False, color="FF0000")
                    else:
                        # 字体格式为红色、宋体、不加粗
                        cell.font = Font(name='宋体', size=11, bold=False, color="008000")

#根据每列中长度最长的单元格设置每个sheet页的列宽，实现列宽自适应
def col_auto_scal(sheet):
    cols = sheet.columns
    col_width = []
    #控制列数
    j = 0
    #获取每列的最长宽度
    for col in cols:
        for i in range(len(col)):
            if i == 0:
                col_width.append(len(str(col[i].value)))
            else:
                if col_width[j] < len(str(col[i].value)):
                    col_width[j] = len(str(col[i].value))
        j += 1
    #对每列的宽度进行设置，实现列宽自适应
    for n in range(len(col_width)):
        col_letter = get_column_letter(n+1)
        #设置单元格列宽---在最大单元格长度基础上加8个字符的长度
        sheet.column_dimensions[col_letter].width = col_width[n] + 8

